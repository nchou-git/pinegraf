from __future__ import annotations

import csv
import json
import logging
import re
import uuid
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

from backend.config import get_settings
from backend.db.store import Store
from backend.enrichment.pdl_client import PdlClient, PdlPersonQuery
from backend.progress import progress_stats

LOGGER = logging.getLogger(__name__)
PDL_PERSON_CONTENT_TYPE = "application/pdl-person+json"


async def run_enrichment(
    source_run_id: uuid.UUID | str,
    seed_file_path: str,
    *,
    store: Store,
    workspace_name: str | None = None,
) -> dict[str, int]:
    settings = get_settings()
    if not settings.pdl_api_key:
        raise RuntimeError("PDL_API_KEY is required for enrichment runs")

    run_id = uuid.UUID(str(source_run_id))
    run = store.get_source_run(run_id)
    if run is None:
        raise ValueError(f"source run not found: {run_id}")

    LOGGER.info(
        "pdl_enrichment_starting run_id=%s seed_file=%s api_key_set=%s",
        run_id,
        seed_file_path,
        bool(settings.pdl_api_key),
    )
    rows = _read_rows(Path(seed_file_path))[: settings.max_pages]
    total = len(rows)
    LOGGER.info("pdl_enrichment_rows_loaded run_id=%s total=%s", run_id, total)
    stats = {"queried": 0, "matched": 0, "no_match": 0, "errors": 0}
    school = workspace_name or settings.workspace_display_name
    pdl = PdlClient(settings.pdl_api_key)

    store.update_source_run(
        run_id,
        stats=progress_stats(
            stats,
            stage="enrich",
            status="running",
            message="Starting PDL enrichment",
            percent=0.0,
        ),
    )

    async with httpx.AsyncClient(timeout=pdl.timeout) as client:
        for index, row in enumerate(rows, start=1):
            first_name = _first_value(row, "first_name")
            last_name = _first_value(row, "last_name")
            class_year = _first_value(row, "class", "class_year")
            url = f"pdl://person/{_slug(first_name)}-{_slug(last_name)}-{_slug(class_year) or 'na'}"
            stats["queried"] += 1

            if not first_name or not last_name:
                stats["errors"] += 1
                store.add_fetch(
                    source_run_id=run_id,
                    url=url,
                    body_bytes=None,
                    http_status=None,
                    error_message="pdl_missing_name",
                )
                _update_progress(store, run_id, stats, index, total)
                continue

            query = PdlPersonQuery(
                first_name=first_name,
                last_name=last_name,
                school=school,
                company=_first_value(
                    row,
                    "current_employer",
                    "current_employer_name",
                    "summer_internship_company_name",
                ),
                location=_first_value(
                    row,
                    "location",
                    "internship_location_city",
                    "state",
                ),
            )
            LOGGER.info(
                "pdl_querying index=%s/%s first=%s last=%s",
                index,
                total,
                first_name,
                last_name,
            )
            try:
                result = await pdl.enrich_person(query, client=client)
            except (httpx.TimeoutException, httpx.RequestError) as exc:
                stats["errors"] += 1
                LOGGER.warning(
                    "pdl_request_failed first=%s last=%s err=%s",
                    first_name,
                    last_name,
                    exc,
                )
                store.add_fetch(
                    source_run_id=run_id,
                    url=url,
                    body_bytes=None,
                    http_status=None,
                    error_message=f"pdl_request_failed: {type(exc).__name__}: {exc}",
                )
                _update_progress(store, run_id, stats, index, total)
                continue
            except Exception as exc:  # noqa: BLE001
                stats["errors"] += 1
                LOGGER.exception("pdl_unexpected_error first=%s last=%s", first_name, last_name)
                store.add_fetch(
                    source_run_id=run_id,
                    url=url,
                    body_bytes=None,
                    http_status=None,
                    error_message=f"pdl_unexpected: {type(exc).__name__}: {exc}",
                )
                _update_progress(store, run_id, stats, index, total)
                continue
            if result.status_code == 200 and result.data:
                LOGGER.info(
                    "pdl_match index=%s/%s likelihood=%s",
                    index,
                    total,
                    result.likelihood,
                )
                stats["matched"] += 1
                body = result.raw_body or json.dumps({"data": result.data}).encode()
                store.add_fetch(
                    source_run_id=run_id,
                    url=url,
                    body_bytes=body,
                    http_status=200,
                    content_type=PDL_PERSON_CONTENT_TYPE,
                    discovery_method=f"pdl:likelihood={result.likelihood}",
                )
            elif result.status_code == 404:
                stats["no_match"] += 1
                store.add_fetch(
                    source_run_id=run_id,
                    url=url,
                    body_bytes=None,
                    http_status=404,
                    error_message="pdl_no_match",
                )
            elif result.status_code == 402:
                stats["errors"] += 1
                store.add_fetch(
                    source_run_id=run_id,
                    url=url,
                    body_bytes=None,
                    http_status=402,
                    error_message="pdl_out_of_credits",
                )
                LOGGER.warning("pdl_out_of_credits source_run_id=%s", run_id)
                _update_progress(store, run_id, stats, index, total)
                break
            else:
                stats["errors"] += 1
                store.add_fetch(
                    source_run_id=run_id,
                    url=url,
                    body_bytes=None,
                    http_status=result.status_code,
                    error_message=result.error or f"http_{result.status_code}",
                )
            _update_progress(store, run_id, stats, index, total)

    cumulative_fetched, cumulative_known = store.refresh_source_crawl_counters(run.source_id)
    store.update_source_run(
        run_id,
        status="complete",
        stats=progress_stats(
            stats,
            stage="enrich",
            status="complete",
            message="Enrichment complete",
            percent=100.0,
            data={"fetched": cumulative_fetched, "known": cumulative_known},
        ),
        finished=True,
    )
    store.mark_source_full_recrawl_complete(run.source_id)
    return stats


def _update_progress(
    store: Store,
    run_id: uuid.UUID,
    stats: dict[str, int],
    index: int,
    total: int,
) -> None:
    store.update_source_run(
        run_id,
        stats=progress_stats(
            stats,
            stage="enrich",
            status="running",
            message=f"Enriching {index} of {total}",
            percent=index / max(total, 1) * 100,
        ),
    )


def _read_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.casefold() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            return [
                {_normal_key(key): value or "" for key, value in row.items()}
                for row in csv.DictReader(handle)
            ]
    if path.suffix.casefold() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_normal_key(value) for value in rows[0]]
        output: list[dict[str, str]] = []
        for raw in rows[1:]:
            output.append(
                {
                    headers[index]: "" if value is None else str(value)
                    for index, value in enumerate(raw)
                    if index < len(headers) and headers[index]
                }
            )
        return output
    raise ValueError(f"unsupported enrichment file type: {path.suffix}")


def _first_value(row: dict[str, Any], *keys: str) -> str:
    normalized = {_normal_key(key): value for key, value in row.items()}
    for key in keys:
        value = normalized.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _normal_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.strip().lower())
    return slug.strip("-")
