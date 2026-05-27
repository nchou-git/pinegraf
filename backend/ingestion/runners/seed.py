from __future__ import annotations

import csv
import re
import uuid
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.config import get_settings
from backend.db.store import Store, content_digest
from backend.ingestion.fetcher import fetch_url
from backend.progress import progress_stats


async def run_seed(
    source_run_id: uuid.UUID | str,
    seed_file_path: str,
    *,
    store: Store,
) -> dict[str, int]:
    run_id = uuid.UUID(str(source_run_id))
    run = store.get_source_run(run_id)
    if run is None:
        raise ValueError(f"source run not found: {run_id}")
    rows = _read_seed_rows(Path(seed_file_path))[: get_settings().max_pages]
    stats = {"queried": 0, "found": 0, "fetched": 0, "missed": 0}
    store.update_source_run(
        run_id,
        stats=progress_stats(
            stats,
            stage="crawl",
            status="running",
            message="Crawling seed rows",
            percent=0.0,
        ),
    )
    total = len(rows)
    for index, row in enumerate(rows, start=1):
        stats["queried"] += 1
        candidates = _candidate_urls(row)
        fetched_for_row = False
        for url in candidates:
            try:
                body = await fetch_url(url, store=store, source_run_id=run_id)
            except Exception as exc:  # noqa: BLE001 - failures are recorded per URL.
                store.add_fetch(
                    source_run_id=run_id,
                    url=url,
                    body_bytes=None,
                    error_message=f"{type(exc).__name__}: {exc}",
                )
                continue
            fetched_for_row = True
            stats["found"] += 1
            stats["fetched"] += 1
            _add_success_fetch_with_hash_diff(
                store,
                source_id=run.source_id,
                source_run_id=run_id,
                url=url,
                body_bytes=body,
                http_status=200,
            )
            break
        if not fetched_for_row:
            stats["missed"] += 1
        if stats["fetched"] % 50 == 0:
            cumulative_fetched, cumulative_known = store.refresh_source_crawl_counters(
                run.source_id
            )
        else:
            current_source = store.get_source(run.source_id)
            cumulative_fetched = (
                int(current_source.pages_fetched_total or 0) if current_source else 0
            )
            cumulative_known = int(current_source.urls_known_total or 0) if current_source else 0
        store.update_source_run(
            run_id,
            stats=progress_stats(
                stats,
                stage="crawl",
                status="running",
                message="Crawling seed rows",
                percent=index / max(total, 1) * 100,
                data={"fetched": cumulative_fetched, "known": cumulative_known},
            ),
        )

    status = "complete" if stats["missed"] == 0 else "partial"
    if stats["fetched"] == 0 and stats["missed"] > 0:
        status = "failed"
    cumulative_fetched, cumulative_known = store.refresh_source_crawl_counters(run.source_id)
    store.update_source_run(
        run_id,
        status=status,
        stats=progress_stats(
            stats,
            stage="crawl",
            status="failed" if status == "failed" else "complete",
            message=status,
            percent=100.0,
            data={"fetched": cumulative_fetched, "known": cumulative_known},
        ),
        finished=True,
    )
    if status == "complete":
        store.mark_source_full_recrawl_complete(run.source_id)
    return stats


def _add_success_fetch_with_hash_diff(
    store: Store,
    *,
    source_id: uuid.UUID,
    source_run_id: uuid.UUID,
    url: str,
    body_bytes: bytes,
    http_status: int,
) -> None:
    digest = content_digest(body_bytes)
    prior = store.latest_successful_fetch_for_url(source_id=source_id, url=url)
    unchanged_since = prior.id if prior is not None and prior.content_hash == digest else None
    store.add_fetch(
        source_run_id=source_run_id,
        url=url,
        body_bytes=None if unchanged_since is not None else body_bytes,
        content_hash=digest,
        body_unchanged_since=unchanged_since,
        http_status=http_status,
    )


def _read_seed_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.casefold() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            return [
                {key: value or "" for key, value in row.items()} for row in csv.DictReader(handle)
            ]
    if path.suffix.casefold() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_normal_key(value) for value in rows[0]]
        output: list[dict[str, str]] = []
        for raw in rows[1:]:
            output.append(
                {
                    headers[index]: "" if value is None else str(value)
                    for index, value in enumerate(raw)
                    if index < len(headers) and headers[index]
                }
            )
        return output
    raise ValueError(f"unsupported seed file type: {path.suffix}")


def _candidate_urls(row: dict[str, Any]) -> list[str]:
    name = _first_value(row, "name", "alum_name", "full_name", "alumni_name")
    if not name:
        return []
    return [f"https://tuck.dartmouth.edu/mba/alumni-stories/{_slug(name)}"]


def _first_value(row: dict[str, Any], *keys: str) -> str:
    normalized = {_normal_key(key): value for key, value in row.items()}
    for key in keys:
        value = normalized.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _normal_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.strip().lower())
    return slug.strip("-")
